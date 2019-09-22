import openpyxl
#import BarChart, Reference, Series

# ------------------------------------------------------------------------

#                          Data Analysis Program
"""Program for analyzing data output from Scouting Application, takes in an
Excel sheet with the data from the Scouting Application organized by auto
and teleop sheets. After changing the columns and placements, it will output
a new Excel sheet that will sort the data by team, calculating averages and 
ranking the teams by the averages * a weight

"""
# ------------------------------------------------------------------------



# ------------------------------------------------------------------------
#                          VARIABLES
# ------------------------------------------------------------------------

# Variables for setting up the Data Analysis Sheet
# Auto Sheet
auto_c1 = "Match Num"
auto_c2 = "Baseline"
auto_c3 = "Cargo Level 1"
auto_c4 = "Cargo Level 2/Cargo ship"
auto_c5 = "Cargo Level 3"
auto_c6 = "Hatch Level 1"
auto_c7 = "Hatch Level 2"
auto_c8 = "Hatch Level 3"
auto_c9 = ""
auto_c10 = ""
auto_c11 = ""
auto_c12 = ""

auto_avg1 = "AVG Cargo"
auto_avg2 = "AVG Hatch"
auto_avg3 = ""

# Tele Sheet
tele_c1 = auto_c1
tele_c2 = auto_c3
tele_c3 = auto_c4
tele_c4 = auto_c5
tele_c5 = auto_c6
tele_c6 = auto_c7
tele_c7 = auto_c8
tele_c8 = "Climb"
tele_c9 = "Notes"
tele_c10 = "Total Cargo"
tele_c11 = "Total Hatch"
tele_c12 = ""
tele_c13 = ""
tele_c14 = ""
tele_c15 = ""

tele_avg1 = "AVG Cargo"
tele_avg2 = "AVG Hatch"
tele_avg3 = "AVG Climb"
tele_avg4 = ""
tele_avg5 = ""

# Overall Team Data
teams_c1 = "Team Num"
teams_c2 = "Auto Cargo AVG"
teams_c3 = "Auto Hatch AVG"
teams_c4 = "Tele Cargo AVG"
teams_c5 = "Tele Hatch AVG"
teams_c6 = "Tele Climb AVG"
teams_c7 = ""
teams_c8 = ""


teams_score = "Score"


# Names of the Excel Sheets
input_data_sheet_name = "2019LVRegionalData" # Name format: <year><Regional>Data
output_analysis_sheet_name = "2019LVRegionalAnalysis0" # Name format: <year><Regional>Analysis

# Weights of each avg
teams_c1_weight = 3
teams_c2_weight = 2
teams_c3_weight = 3
teams_c4_weight = 2
teams_c5_weight = 1
teams_c6_weight = 0
teams_c7_weight = 0
teams_c8_weight = 0



# ------------------------------------------------------------------------
#                          METHODS
# -----------------------------------------------------------------------
def sheet_setup(sheet, teamnum):
    # Method sets up sheet for each team

    # set top of sheet to say team number
    sheet.cell(1,5).value = teamnum

    # --------------------------------------------------
    #         AUTO
    # --------------------------------------------------
    # col 1 - match num
    # col 2 - baseline
    # col 3 - cargo 1
    # col 4 - cargo 2
    # col 5 - cargo 3
    # col 6 - hatch 1
    # col 7 - hatch 2
    # col 8 - hatch 3

    # col 10 - avg cargo
    # col 11 - avg hatch


    sheet.cell(2,2).value = "AUTO"
    # AUTO

    sheet.cell(3, 1).value = auto_c1
    sheet.cell(3, 2).value = auto_c2
    sheet.cell(3, 3).value = auto_c3
    sheet.cell(3, 4).value = auto_c4
    sheet.cell(3, 5).value = auto_c5
    sheet.cell(3, 6).value = auto_c6
    sheet.cell(3, 7).value = auto_c7
    sheet.cell(3, 8).value = auto_c8
    sheet.cell(3, 9).value = auto_c9
    sheet.cell(3, 10).value = auto_c10
    sheet.cell(3, 11).value = auto_c11
    sheet.cell(3, 12).value = auto_c12

    sheet.cell(3, 13).value = auto_avg1
    sheet.cell(3, 14).value = auto_avg2
    sheet.cell(3, 15).value = auto_avg3


    # ----------------------------------------------------------
    #                     TELEOP
    # ----------------------------------------------------------
    # col 1 - match num
    # col 2 - Balls low
    # col 3 - Balls mid
    # col 4 - balls high
    # col 5 - hatch low
    # col 6 - hatch mid
    # col 7 - hatch high
    # col 8 - climb
    # col 9 - notes
    # col 10 - total cargo
    # col 11 - total hatch

    # col 13 - avg switch
    # col 14 - avg vault
    # col 15 - avg scale
    # input is the Excel sheet
    # TELEOP
    sheet.cell(24, 2).value = "TELEOP"

    sheet.cell(25, 1).value = tele_c1
    sheet.cell(25, 2).value = tele_c2
    sheet.cell(25, 3).value = tele_c3
    sheet.cell(25, 4).value = tele_c4
    sheet.cell(25, 5).value = tele_c5
    sheet.cell(25, 6).value = tele_c6
    sheet.cell(25, 7).value = tele_c7
    sheet.cell(25, 8).value = tele_c8
    sheet.cell(25, 9).value = tele_c9
    sheet.cell(25, 10).value = tele_c10
    sheet.cell(25, 11).value = tele_c11
    sheet.cell(25, 12).value = tele_c12
    sheet.cell(25, 13).value = tele_c13
    sheet.cell(25, 14).value = tele_c14
    sheet.cell(25, 15).value = tele_c15

    sheet.cell(25, 16).value = tele_avg1
    sheet.cell(25, 17).value = tele_avg2
    sheet.cell(25, 18).value = tele_avg3
    sheet.cell(25, 19).value = tele_avg4
    sheet.cell(25, 20).value = tele_avg5


# --------------------------------------------------------------------------------------

def transfer_tele_data(data_sheet, output_sheet, d_row, o_row):
    # transfer tele data from data sheet to the output sheet for each team
    # in data excel sheet
    # col 2 - team num
    # col 3 - match num
    # col 4 - cubes in switch
    # col 5 - cubes in vault
    # col 6 - cubes in scale
    # col 7 - Climb (Yes/No)
    # col 8 - Result (Win/Lost/Tie)

    # OUTPUT SHEET
    # col 1 - match num
    # col 2 - Balls low
    # col 3 - Balls mid
    # col 4 - balls high
    # col 5 - hatch low
    # col 6 - hatch mid
    # col 7 - hatch high
    # col 8 - climb
    # col 9 - notes

    # col 11 - avg switch
    # col 12 - avg vault
    # col 13 - avg scale

    #   set MATCH NUM
    output_sheet.cell(o_row, 1).value = data_sheet.cell(d_row, 3).value
#   set CARGO 1
    output_sheet.cell(o_row, 2).value = data_sheet.cell(d_row, 4).value
#   set CARGO 2
    output_sheet.cell(o_row, 3).value = data_sheet.cell(d_row, 5).value
#   set CARGO 3
    output_sheet.cell(o_row, 4).value = data_sheet.cell(d_row, 6).value
#   set HATCH 1
    output_sheet.cell(o_row, 5).value = data_sheet.cell(d_row, 7).value
#   set HATCH 2
    output_sheet.cell(o_row, 6).value = data_sheet.cell(d_row, 8).value
#   set HATCH 3
    output_sheet.cell(o_row, 7).value = data_sheet.cell(d_row, 9).value
#   set CLIMB
    output_sheet.cell(o_row, 8).value = data_sheet.cell(d_row, 10).value

#   set NOTES
    output_sheet.cell(o_row, 9).value = data_sheet.cell(d_row, 11).value
    #   set TOTAL CARGO

    output_sheet.cell(o_row, 10).value = int(output_sheet.cell(o_row, 2).value) + int(output_sheet.cell(o_row, 3).value) \
      + int(output_sheet.cell(o_row, 4).value)

#   set TOTAL HATCH
    output_sheet.cell(o_row, 11).value = int(output_sheet.cell(o_row, 5).value) + int(output_sheet.cell(o_row, 6).value) \
        + int(output_sheet.cell(o_row, 7).value)
    print(o_row)
# -----------------------------------------------------------------------

def transfer_auto_data(data_sheet, output_sheet, d_row, o_row):
    # transfer auto data from data sheet to the output sheet for each team
    # in data excel sheet
    # col 2 - team num
    # col 3 - match num
    # col 4 - cubes in switch
    # col 5 - baseline
    # col 6 - cubes in scale

    # OUTPUT SHEET
    # col 1 - match num
    # col 2 - baseline
    # col 3 - cargo 1
    # col 4 - cargo 2
    # col 5 - cargo 3
    # col 6 - hatch 1
    # col 7 - hatch 2
    # col 8 - hatch 3

    # col 10 - avg cargo
    # col 11 - avg hatch

    #   set MATCH NUM
    output_sheet.cell(o_row, 1).value = data_sheet.cell(d_row, 3).value
#   set BASELINE
    output_sheet.cell(o_row, 2).value = data_sheet.cell(d_row, 4).value
#   set CARGO 1
    output_sheet.cell(o_row, 3).value = data_sheet.cell(d_row, 5).value
#   set CARGO 2
    output_sheet.cell(o_row, 4).value = data_sheet.cell(d_row, 6).value
#   set CARGO 3
    output_sheet.cell(o_row, 5).value = data_sheet.cell(d_row, 7).value
#   set HATCH 1
    output_sheet.cell(o_row, 6).value = data_sheet.cell(d_row, 8).value
#   set HATCH 2
    output_sheet.cell(o_row, 7).value = data_sheet.cell(d_row, 9).value
#   set HATCH 3
    output_sheet.cell(o_row, 8).value = data_sheet.cell(d_row, 10).value

#   set TOTAL
    output_sheet.cell(o_row, 4).value = int(output_sheet.cell(o_row, 3).value) + int(output_sheet.cell(o_row, 2).value)

def print_team_stat(overall_team_sheet, row, team, stat_list):
    # 1) = "Team Num"
    # 2) = "Auto cargo AVG"
    # 3) = "Auto hatch AVG"
    # 4) = "Tele cargo AVG"
    # 5) = "Tele hatch AVG"
    # 6) = "Tele climb AVG"
    # 7) = "Score"
    overall_team_sheet.cell(row, 1).value = team
    overall_team_sheet.cell(row, 2).value = stat_list[0]
    overall_team_sheet.cell(row, 3).value = stat_list[1]
    overall_team_sheet.cell(row, 4).value = stat_list[2]
    overall_team_sheet.cell(row, 5).value = stat_list[3]
    overall_team_sheet.cell(row, 6).value = stat_list[4]

    # weight factors of each part that makes up the score. Can be adjusted
    # ----------------------------------------------------------------------------------------
    a_cargo_factor = teams_c1_weight
    a_hatch_factor = teams_c2_weight
    t_cargo_factor = teams_c3_weight                   # <<<<<<   ADJUST WEIGH FACTOR
    t_hatch_factor = teams_c4_weight
    t_climb_factor = teams_c5_weight
    # ----------------------------------------------------------------------------------------

    score = a_cargo_factor * stat_list[0] + a_hatch_factor * stat_list[1] + \
            t_cargo_factor * stat_list[2] + t_hatch_factor * stat_list[3] + t_climb_factor * stat_list[4]


    overall_team_sheet.cell(row, 9).value = score


def calc_avg(output_sheet, matches_played):
    '''
    Calculates both averages from auto and teleop data
    :param output_sheet: team sheet
    :param matches_played: how long the row goes for
    :param col:
    :return: prints the averages to the sheet
    '''

    # AUTO
    # ----------------------------
    # average switch
    a_cargo_sum = 0
    a_hatch_sum = 0
  # finds sum of all the switch and scale cubes
    for i in range(4, 4 + matches_played, 1):

        a_cargo_sum += int(output_sheet.cell(i,3).value) + int(output_sheet.cell(i,4).value) + int(output_sheet.cell(i,5).value)
        a_hatch_sum += output_sheet.cell(i,6).value + output_sheet.cell(i,7).value + output_sheet.cell(i,8).value
# prints out average to the Excel sheet
    if matches_played == 0:
        output_sheet.cell(4, 13).value = 0
        output_sheet.cell(4, 14).value = 0
    else:
        output_sheet.cell(4, 13).value = float(a_cargo_sum)/matches_played
        output_sheet.cell(4, 14).value = float(a_hatch_sum)/matches_played


    # TELE
    # ------------------------------
    # OUTPUT SHEET
    # col 1 - match num
    # col 2 - Balls low
    # col 3 - Balls mid
    # col 4 - balls high
    # col 5 - hatch low
    # col 6 - hatch mid
    # col 7 - hatch high
    # col 8 - climb
    # col 9 - notes

    # col 11 - avg switch
    # col 12 - avg vault
    # col 13 - avg scale
    t_cargo_sum = 0
    t_hatch_sum = 0
    t_climb_sum = 0

    for i in range(26, 26 + matches_played, 1):
        t_cargo_sum += output_sheet.cell(i, 2).value + output_sheet.cell(i, 3).value + output_sheet.cell(i, 4).value
        t_hatch_sum += output_sheet.cell(i, 5).value + output_sheet.cell(i, 6).value + output_sheet.cell(i, 7).value

        if output_sheet.cell(i, 8).value == 1:
            t_climb_sum += 3
        elif output_sheet.cell(i, 8).value == 2:
            t_climb_sum += 6
        elif output_sheet.cell(i, 8).value == 3:
            t_climb_sum += 12
        else:
            t_climb_sum += 0

    if matches_played == 0:
        output_sheet.cell(26, 16).value = 0
        output_sheet.cell(26, 17).value = 0
        output_sheet.cell(26, 18).value = 0
    else:
        output_sheet.cell(26, 16).value = float(t_cargo_sum)/matches_played
        output_sheet.cell(26, 17).value = float(t_hatch_sum)/matches_played
        output_sheet.cell(26, 18).value = float(t_climb_sum)/matches_played

    return [output_sheet.cell(4, 13).value, output_sheet.cell(4,14).value,
            output_sheet.cell(26, 16).value, output_sheet.cell(26, 17).value , output_sheet.cell(26, 18).value]







#

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
#                                               MAIN CODE
# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------






# instantiate new variables
# data is the excel sheet with all the scouting data


# -----------------------------------------------------------------------------------
                                 #   INPUT FILE NAME
data = openpyxl.load_workbook(input_data_sheet_name + '.xlsx')           # <<< Change Name of Input File

# -----------------------------------------------------------------------------------
# auto sheet of the scouting data Excel file
auto = data['Auto']
# tele sheet of the scouting data Excel file
tele = data['Tele']




# Creating new workbook to store organized data
teamOutput = openpyxl.Workbook()

# -----------------------------------------------------------------------------------
                                #    OUTPUT FILE NAME
teamOutput_name = output_analysis_sheet_name + '.xlsx'             # <<< Change Name of Output File

# -----------------------------------------------------------------------------------


teamSheet = teamOutput.active
teamSheet.title = "Overall Team"

# OVERALL TEAM SHEET SET UP
teamSheet.cell(1,4).value = "Overall Team Data"
    # AUTO

teamSheet.cell(2, 1).value = teams_c1
teamSheet.cell(2, 2).value = teams_c2
teamSheet.cell(2, 3).value = teams_c3
teamSheet.cell(2, 4).value = teams_c4
teamSheet.cell(2, 5).value = teams_c5
teamSheet.cell(2, 6).value = teams_c6
teamSheet.cell(2, 7).value = teams_c7
teamSheet.cell(2, 8).value = teams_c8
teamSheet.cell(2, 9).value = teams_score



# ------------------------
#          TELEOP
# ------------------------
# in data excel sheet
# col 2 - team num
# col 3 - match num
# col 4 - Balls low
# col 5 - Balls mid
# col 6 - balls high
# col 7 - hatch low
# col 8 - hatch mid
# col 9 - hatch high
# col 10 - climb
# col 11 - notes


# --------------------
# Set up First team
# --------------------
# set lastTeam to the first team to start
lastTeam = tele.cell(4,2).value

teamOutput.create_sheet(str(lastTeam))
teamSheet = teamOutput[str(lastTeam)]  # set the first team sheet title to the first team
sheet_setup(teamSheet, lastTeam)

# ask for number of number of matches
print "How many data points is there?"
matches = int(raw_input("Data Points:"))

played = 0

team_count = 0
# data to keep overall track of
teams_cargo_avg = []
teams_hatch_avg = []
teams_climb_avg = []
# format of output sheet
# col 1 - match num
# col 2 - switch
# col 3 - vault
# col 4 - scale
# col 5 - climb
# col 6 - result
# col 8 - avg switch
# col 9 - avg vault
# col 10 - avg scale

for i in range(4, matches + 2, 1): # row 4 is start of data
    # ------------------------------------------------------------------
    # SAME TEAM NUMBER
    if lastTeam == tele.cell(i,2).value: #  if same team num add to sum counter
        transfer_tele_data(tele, teamSheet, i, 26 + played)
        transfer_auto_data(auto, teamSheet, i, 4 + played)
        played += 1
        # Set match num
        # ROW starts at 3
        # auto starts at 4
        # teleop starts at 21






    # -------------------------------------------------------------------
    # NEW TEAM NUMBER
    else:
        # calculate the averages of the previous team
        team_stat_list = calc_avg(teamSheet, played)
        # start at row 3 for overall team data
        print_team_stat(teamOutput["Overall Team"], 3 + team_count, lastTeam, team_stat_list)

        teamSheet.cell(1,1).value = played

        #  start of different team num
        lastTeam = tele.cell(i,2).value
        teamOutput.create_sheet(str(lastTeam))
        teamSheet = teamOutput[str(lastTeam)]
        sheet_setup(teamSheet, lastTeam)

        # need to transfer data for first tele value
        transfer_auto_data(auto, teamSheet, i, 4)
        transfer_tele_data(tele, teamSheet, i, 26)
        played = 1
        # increment team count
        team_count += 1


teamOutput.save(teamOutput_name)



